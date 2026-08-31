import socket
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime

# ── 설정 ──
UDP_PORT    = 12345
BUFFER_SIZE = 1024

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
FILENAME  = f"C:/Users/mugi4/Desktop/가천대/공모전/화물지킴이_실험데이터_{timestamp}.xlsx"

# ── 색상 정의 ──
FILL_WARNING = PatternFill("solid", fgColor="FFA500")  # 주황
FILL_DANGER  = PatternFill("solid", fgColor="FF0000")  # 빨강
FILL_RESET   = PatternFill("solid", fgColor="00CC00")  # 초록
FONT_WHITE   = Font(color="FFFFFF", bold=True)
FONT_BLACK   = Font(color="000000", bold=False)

# ── 엑셀 초기화 ──
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "실험데이터"

# ★ v3: 자이로 1~4번 개별 % 컬럼 추가 (기존 13개 → 17개)
headers = [
    "시간(ms)",
    "기울기raw(최대값,°)", "장력raw", "거리raw(mm)",
    "기울기(%,대표값)", "장력(%)", "거리(%)",
    "자이로점수", "장력점수", "거리점수",
    "총점", "최종단계", "경보상태", "기록시각",
    "자이로1(%)", "자이로2(%)", "자이로3(%)", "자이로4(%)"
]
ws.append(headers)

header_fill = PatternFill("solid", fgColor="1F4E79")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = Font(color="FFFFFF", bold=True)

col_widths = [12, 16, 12, 14, 14, 10, 10, 12, 12, 12, 10, 12, 12, 20, 12, 12, 12, 12]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

# ── UDP 소켓 ──
sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
sock.bind(("0.0.0.0", UDP_PORT))
sock.settimeout(1.0)

print("수신 대기 중...")
print(f"저장 파일: {FILENAME}")
print("종료: Ctrl+C")
print("=" * 50)

row = 2

try:
    while True:
        try:
            data, addr = sock.recvfrom(BUFFER_SIZE)
            line = data.decode("utf-8", errors="ignore").strip()
        except socket.timeout:
            continue

        if not line:
            continue

        # ── 리셋 처리 ──
        if line == "RESET":
            now = datetime.now().strftime("%H:%M:%S")
            ws.append(["===RESET==="] + [""] * 16 + [now])
            for col in range(1, 19):
                ws.cell(row=row, column=col).fill = FILL_RESET
                ws.cell(row=row, column=col).font = FONT_WHITE
            row += 1
            wb.save(FILENAME)
            print(f"[{now}] ===== 영점 리셋 =====")
            continue

        if not line.startswith("CSV,"):
            continue

        parts = line.split(",")
        parts = parts[1:]

        # ★ v3: 13개 → 17개 필드로 변경
        if len(parts) != 17:
            continue

        try:
            time_ms     = int(parts[0])
            raw_tilt    = float(parts[1])
            raw_tension = int(parts[2])
            raw_dist    = int(parts[3])
            pct_tilt    = float(parts[4])
            pct_tension = float(parts[5])
            pct_dist    = float(parts[6])
            score_t     = int(parts[7])
            score_f     = int(parts[8])
            score_d     = int(parts[9])
            total_score = int(parts[10])
            level       = int(parts[11])
            alarm       = parts[12]
            gyro1       = float(parts[13])
            gyro2       = float(parts[14])
            gyro3       = float(parts[15])
            gyro4       = float(parts[16])
            now         = datetime.now().strftime("%H:%M:%S")

            ws.append([
                time_ms,
                raw_tilt, raw_tension, raw_dist,
                pct_tilt, pct_tension, pct_dist,
                score_t, score_f, score_d,
                total_score, level, alarm, now,
                gyro1, gyro2, gyro3, gyro4
            ])

            # ── 경보 상태 열 색칠 (13번) ──
            if alarm == "DANGER":
                ws.cell(row=row, column=13).fill = FILL_DANGER
                ws.cell(row=row, column=13).font = FONT_WHITE
            elif alarm == "WARNING":
                ws.cell(row=row, column=13).fill = FILL_WARNING
                ws.cell(row=row, column=13).font = FONT_BLACK

            # ── 총점 열 색칠 (11번) ──
            if total_score >= 6:
                ws.cell(row=row, column=11).fill = FILL_DANGER
                ws.cell(row=row, column=11).font = FONT_WHITE
            elif total_score >= 3:
                ws.cell(row=row, column=11).fill = FILL_WARNING

            # ── 센서 점수 열 색칠 (8,9,10번) ──
            score_cols = [score_t, score_f, score_d]
            for i, score in enumerate(score_cols):
                col = 8 + i
                if score >= 2:
                    ws.cell(row=row, column=col).fill = FILL_DANGER
                    ws.cell(row=row, column=col).font = FONT_WHITE
                elif score >= 1:
                    ws.cell(row=row, column=col).fill = FILL_WARNING

            # ── 대표 퍼센트 열 색칠 (5,6,7번) ──
            pct_vals = [pct_tilt, pct_tension, pct_dist]
            pct_minors = [300.0, 50.0, 15.0]   # Arduino 코드의 MINOR 임계값(%) 기준
            pct_majors = [400.0, 80.0, 30.0]   # Arduino 코드의 MAJOR 임계값(%) 기준
            for i, (pct, minor, major) in enumerate(zip(pct_vals, pct_minors, pct_majors)):
                col = 5 + i
                if pct >= major:
                    ws.cell(row=row, column=col).fill = FILL_DANGER
                    ws.cell(row=row, column=col).font = FONT_WHITE
                elif pct >= minor:
                    ws.cell(row=row, column=col).fill = FILL_WARNING

            # ── 자이로 1~4번 개별 열 색칠 (15,16,17,18번) ──
            gyro_vals = [gyro1, gyro2, gyro3, gyro4]
            for i, g in enumerate(gyro_vals):
                col = 15 + i
                if g >= 400.0:
                    ws.cell(row=row, column=col).fill = FILL_DANGER
                    ws.cell(row=row, column=col).font = FONT_WHITE
                elif g >= 300.0:
                    ws.cell(row=row, column=col).fill = FILL_WARNING

            row += 1

            print(f"[{now}] "
                  f"기울기(대표):{raw_tilt:.1f}°({pct_tilt:.1f}%/{score_t}pt) "
                  f"[T1:{gyro1:.0f}% T2:{gyro2:.0f}% T3:{gyro3:.0f}% T4:{gyro4:.0f}%] "
                  f"장력:{raw_tension}({pct_tension:.1f}%/{score_f}pt) "
                  f"거리:{raw_dist}mm({pct_dist:.1f}%/{score_d}pt) "
                  f"총점:{total_score}pt → {alarm}")

            if row % 100 == 0:
                wb.save(FILENAME)
                print(f"자동 저장! ({row}행)")

        except ValueError:
            continue

except KeyboardInterrupt:
    print("\n종료 중...")
finally:
    try:
        wb.save(FILENAME)
        print(f"저장 완료! → {FILENAME}")
    except Exception as e:
        print(f"저장 실패: {e}")
    finally:
        sock.close()